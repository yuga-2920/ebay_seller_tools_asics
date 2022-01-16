from collections import OrderedDict
import eel
import pandas as pd
import re
import requests
from googletrans import Translator
import openpyxl
from bs4 import BeautifulSoup
import traceback

from change_description import change_description

# 靴のサイズの変更
size_dict = {
    '22.5':'US4/JP22.5',
    '23':'US4.5/JP23.0',
    '23.5':'US5/JP23.5',
    '24':'US5.5/JP24.0',
    '24.5':'US6/JP24.5',
    '25':'US6.5/JP25',
    '25.5':'US7.5/JP25.5',
    '26':'US8/JP26',
    '26.5':'US8.5/JP26.5',
    '27':'US9/JP27',
    '27.5':'US9.5/JP27.5',
    '28':'US10/JP28',
    '28.5':'US11/JP28.5',
    '29':'US11.5/JP29',
    '29.5':'US12/JP29.5',
    '30':'US12.5/JP30',
    '30.5':'US13/JP30.5',
    '31':'US14/JP31',
    '32':'US15/JP32',
    '33':'US16/JP33'
}

class MyScraping():

    def __init__(self, items_title, items_size, items_id, items_price, items_image_url, items_description, items_name, items_part_num):
        self.items_title = items_title
        self.items_size = items_size
        self.items_id = items_id
        self.items_price = items_price
        self.items_image_url = items_image_url
        self.items_description = items_description
        self.items_name = items_name
        self.items_part_num = items_part_num

    #翻訳関数
    def translate(self, string):
        translator = Translator()
        result = translator.translate(string, src='ja', dest='en')
        return result.text

    #翻訳前の不要ワードの設定
    def delete_word_before(self, word, delete_before_list):

        code_regex = re.compile('[!"#$&*+-.:;<=>?@^_`{|}~「」〔〕“”〈〉『』＆＊・＄＃＠。、？！｀＋￥◆★◇☆⇒%％◎○■/:×,[]【】/,()★:❣❣#]')
        word = code_regex.sub(" ", word)

        for delete_word in delete_before_list:
            word = word.replace(delete_word, "")

        return word

    #翻訳後の不要ワードの設定
    def delete_word_after(self, word, delete_after_list, add_name_list):

        word_list = word.split(" ")

        for delete_word in delete_after_list:
            if delete_word.lower() in word_list:
                word_list.remove(delete_word.lower())
            if delete_word.capitalize() in word_list:
                word_list.remove(delete_word.capitalize())

        for add_word in add_name_list:
            if add_word.lower() in word_list:
                word_list.remove(add_word.lower())
            if add_word.capitalize() in word_list:
                word_list.remove(add_word.capitalize())

        return word_list

    #ワードの追加
    def add_word(self, word_num, item_name, item_color, item_part_num, add_word_list):
        
        title_len = len(item_name) + len(item_color) + len(item_part_num)
        exlude_name_len = len(item_color) + len(item_part_num)

        if title_len == word_num:
            return item_name
        
        elif title_len > word_num:
            while exlude_name_len + len(item_name) > word_num:
                word_list = item_name.split(' ')
                item_name = ' ' + ' '.join(word_list[:-1])

                if item_name == " " or item_name == "  ":
                    return ""

            if exlude_name_len + len(item_name)  == word_num:
                return item_name
            
            else:
                item_name = item_name + ' '
                for add_word in add_word_list:
                    if add_word in item_name:
                        continue

                    if exlude_name_len + len(item_name) + len(add_word) >= word_num:
                        return item_name

                    item_name = item_name  + item_name + " "

                else:
                    return item_name

        elif title_len < word_num:
            item_name = item_name + " "
            for add_word in add_word_list:
                if add_word in item_name:
                    continue

                if exlude_name_len + len(item_name) + len(add_word) >= word_num:
                    return item_name

                item_name = item_name  + add_word + " "

            else:
                return item_name

    #商品名のチェック
    def check_item_name(self, item_name, exlude_word_list, in_word_list):

        for exlude_word in exlude_word_list:
            
            if exlude_word != '' and exlude_word in item_name:
                return ''

        for in_word in in_word_list:
            if in_word != '' and in_word not in item_name:
                return ''

        return item_name

    #商品名の変更
    def change_item_name(self, item_name, item_color, item_part_num, kataban_check_flug, add_name_len, add_word_list, delete_after_list, delete_before_list, add_name_split_dot):

        if kataban_check_flug:
            
            # 英数字もしくはハイフンを抽出
            item_name_kataban = re.findall('[a-zａ-ｚA-ZＡ-Ｚ0-9０-９_]+|-', item_name, flags=re.IGNORECASE)
            item_name = ' '.join(item_name_kataban)

            item_name_delete_list = self.delete_word_after(item_name, delete_after_list, add_name_split_dot)
            item_name = ' '.join(OrderedDict.fromkeys(item_name_delete_list))
            item_name = item_name.upper()

            item_name_result = self.add_word(add_name_len, item_name, item_color, item_part_num, add_word_list)
            item_name = add_name_split_dot[0] + ' ' + item_name_result + add_name_split_dot[1]

        else:
            item_name_ja_delete = self.delete_word_before(item_name, delete_before_list)
            item_name = self.translate(item_name_ja_delete)
            item_name = item_name.upper()

            item_name_delete_list = self.delete_word_after(item_name, delete_after_list, add_name_split_dot)
            item_name = ' '.join(OrderedDict.fromkeys(item_name_delete_list))

            item_name_result = self.add_word(add_name_len, item_name, item_color, item_part_num, add_word_list)
            item_name = add_name_split_dot[0] + ' ' + item_name_result + add_name_split_dot[1]

        return item_name

    # 靴のサイズの変更
    def change_shoose_size(self, shoose_size):
        return size_dict[shoose_size]

    # listの大きさをそろえる
    def create_same_size_list(self, size_list, list, append_word):
        return list + [append_word] * (len(size_list) - len(list))


    #一覧ページから各商品のURLを取ってくる
    def get_item_url(self, url_eel, eel_page, exlude_word_eel, in_word_eel):
        url_list = []
        page_flag = True

        #商品名の除外ワード、指定ワード、ページ数の読み込み
        num_list = eel_page.split(',')
        exlude_word_list = exlude_word_eel.split(',')
        in_word_list = in_word_eel.split(',')


        #ページごと
        for i in range(int(num_list[0]), int(num_list[1]) + 1):
            eel.view_log_js(str(i) + "ページ目")

            if page_flag:
                page_url = url_eel + "&start=" + str((i - 1) * 24)

                res = requests.get(page_url)
                soup = BeautifulSoup(res.text, 'lxml')

                flag = soup.find(id = "search-result-items")

                if flag == None:
                    page_flag = False

                    page_url = url_eel + "?start=" + str((i - 1) * 24)
                    res = requests.get(page_url)
                    soup = BeautifulSoup(res.text, 'lxml')

            else:
                page_url = url_eel + "?start=" + str((i - 1) * 24)

                res = requests.get(page_url)
                soup = BeautifulSoup(res.text, 'lxml')

            item_list = soup.find_all(class_ = "product-tile__link")
            for j, item in enumerate(item_list):
                item_name = item.find(class_ = "product-tile__text--underline").get_text(strip=True)
                item_name = self.check_item_name(item_name, exlude_word_list, in_word_list)

                if item_name != "":
                    url_list.append(item["href"])

                eel.view_log_js(str(j + 1) + "/" + str(len(item_list)) + " 商品目")

            else:
                #重複の削除
                url_list = list(dict.fromkeys(url_list))

        return url_list

    #一覧ページから情報を取り出す
    def get_item_detail(self, url_list, eel_delete_word_before, eel_delete_word_after, eel_fill_in_word, eel_add_word):
        #ファイルの読み込み
        delete_before_file = pd.read_csv('./' + eel_delete_word_before, header=None, names=['削除するワード_before'])
        delete_before_list = delete_before_file['削除するワード_before'].tolist()
        delete_after_file = pd.read_csv('./' + eel_delete_word_after, header=None, names=['削除するワード_after'])
        delete_after_list = delete_after_file['削除するワード_after'].tolist()
        add_word_file = pd.read_csv('./' + eel_fill_in_word, header=None, names=['追加するワード'])
        add_word_list = add_word_file['追加するワード'].tolist()

        #型番のみの抽出か判別するフラグ
        kataban_check_flug = eel.kataban_check()()

        #商品名の処理
        add_name_len = 80 - len(eel_add_word) - 2
        add_name_split_dot = eel_add_word.split(',')

        #各商品ごと
        for i, top_item_url in enumerate(url_list):
            eel.view_log_js(str(i + 1) + "/" + str(len(url_list)) + " 商品名")
            
            try:
                res = requests.get(top_item_url)
                soup = BeautifulSoup(res.text, 'lxml')

                # カラー毎の抽出
                colors = soup.find_all(class_ = "variants__item--color")
                for j, color in enumerate(colors):
                    eel.view_log_js(str(j + 1) + "/" + str(len(colors)) + " カラー目")
                    if j + 1 == len(colors) + 1:
                        eel.view_log_js("\n")

                    a_tag = color.find("a")
                    url = a_tag["href"]

                    # 商品IDの取得
                    item_id = url.split('/p/')[1].split("?")[0]

                    if item_id in self.items_id:
                        continue

                    res = requests.get(url)
                    soup = BeautifulSoup(res.text, 'lxml')

                    # カラー名の取得
                    item_color = soup.find(class_ = "variants__header--light").get_text(strip=True)
                    item_color = self.translate(item_color).replace(" /", "/").replace("/ ", "/")

                    # 品番
                    item_part_num = soup.find(class_ = "product-number")
                    item_part_num = item_part_num.find_all("span")[1].get_text(strip=True).split(".")[0]
                    print(item_part_num)

                    # 商品名の取得
                    item_name = soup.find("h1", itemprop="name").get_text(strip=True) + " "

                    change_item_name = self.change_item_name(item_name, item_color, item_part_num, kataban_check_flug, add_name_len, add_word_list, delete_after_list, delete_before_list, add_name_split_dot)

                    # 出力タイトル
                    item_title = change_item_name + " " + item_color + " " + item_part_num
                    print(item_title)
                    print(len(item_title))

                    # 価格
                    item_price = soup.find(class_ = "price-sales").get_text(strip=True).replace("セール料金", "").replace("¥", "").replace(",", "").replace("(税込)", "").replace(" ", "")
                    item_price = int(item_price)

                    # 靴のサイズ
                    all_size_list = []
                    item_size_append_list = []
                    size_list = soup.find_all(class_ = "variants__item--size")
                    for size in size_list:
                        if size["data-instock"] == "true":
                            size = self.change_shoose_size(size.get_text(strip=True))
                            all_size_list.append(size)

                    all_size = "Size=" + ",".join(all_size_list)
                    if all_size == "Size=":
                        continue

                    item_size_append_list.append(all_size)

                    for size in all_size_list:
                        if size != "":
                            item_size_append_list.append("Size=" + str(size))

                    # 画像の取得
                    image_append_list = []
                    image_list = soup.find_all(class_ = "thumbnail-link")
                    if len(image_list) >= 2:
                        image_list[0], image_list[1] = image_list[1], image_list[0]
                    for image in image_list:
                        if image["href"] == "#":
                            continue
                        image_append_list.append(image["href"])
                        if len(image_append_list) == 10:
                            break

                    images_url = "|".join(image_append_list)

                    # 詳細の取得
                    origin_description = soup.find(class_ = "product-info-section-inner")
                    p_tag_item_description = origin_description.find("p").get_text("<br>", strip=True)

                    tr_tag_item_description = origin_description.find_all("tr")
                    tr_tag_item_description_list = []
                    for item in tr_tag_item_description:
                        if item != "":
                            tr_tag_item_description_list.append(item.get_text(" ", strip=True))

                    on_item_description = self.translate(p_tag_item_description)
                    
                    under_item_description = "<br>".join(tr_tag_item_description_list)
                    under_item_description = self.translate(under_item_description)
                    under_item_description = under_item_description.replace("<br> <br>", "<br>")

                    item_description = on_item_description + "<br><br>" + under_item_description
                    item_description = change_description(item_description)

                    #listの格納
                    self.items_size.extend(item_size_append_list)
                    self.items_title.append(item_title)
                    self.items_id.append(item_id)
                    self.items_price.append("")
                    self.items_image_url.append(images_url)
                    self.items_description.append(item_description)
                    self.items_name.append(item_name)
                    self.items_part_num.append(item_part_num)

                    # listの大きさを同じにする
                    self.items_title = self.create_same_size_list(self.items_size, self.items_title, "")
                    self.items_id = self.create_same_size_list(self.items_size, self.items_id, "")
                    self.items_price = self.create_same_size_list(self.items_size, self.items_price, item_price)
                    self.items_image_url = self.create_same_size_list(self.items_size, self.items_image_url, "")
                    self.items_description = self.create_same_size_list(self.items_size, self.items_description, "")
                    self.items_name = self.create_same_size_list(self.items_size, self.items_name, "")
                    self.items_part_num = self.create_same_size_list(self.items_size, self.items_part_num, "")

            except KeyboardInterrupt:
                eel.view_log_js("プログラムを強制的に終了しました")
                return self.items_title, self.items_size, self.items_id, self.items_price, self.items_image_url, self.items_description, self.items_name, self.items_part_num

            except Exception:
                t = traceback.format_exc()
                eel.view_log_js(t)
                eel.view_log_js("予期せぬエラーが発生しました。")
                continue

        return self.items_title, self.items_size, self.items_id, self.items_price, self.items_image_url, self.items_description, self.items_name, self.items_part_num

def main():

    try:

        myscraping = MyScraping(items_title=[], items_size=[], items_id=[], items_price=[], items_image_url=[], items_description=[], items_name=[], items_part_num=[])

        eel.view_log_js('\n商品一覧ページから各商品の情報を取得します\n')

        url_list = myscraping.get_item_url(eel.url()(), eel.page()(), eel.exclude_word()(), eel.in_word()())

        eel.view_log_js('\n商品一覧ページから各商品の情報の取得が終わりました')


        eel.view_log_js('\n\n各商品の詳細情報を取得します\n')

        items_title, items_size, items_id, items_price, items_image_url, items_description, items_name, items_part_num = myscraping.get_item_detail(url_list, eel.delete_word_before()(), eel.delete_word_after()(), eel.fill_in_word()(), eel.add_word()())

        eel.view_log_js('\n各商品の詳細情報の取得が終わりました。')

    finally:
        eel.view_log_js('\nファイルの作成をします')

        result_dic = {
            "E": items_title,
            "G": items_size,
            "H": items_id,
            "I": items_price,
            "J": items_image_url,
            "K": items_description,
            "V": items_name,
            "AK": items_part_num
            }

        wb = openpyxl.Workbook()
        ws = wb.active

        ws["E1"] = "*Title"
        ws["G1"] = "RelationshipDetails"
        ws["H1"] = "CustomLabel"
        ws["I1"] = "送料含めた価格"
        ws["J1"] = "PicURL"
        ws["K1"] = "*Description"
        ws["V1"] = "*C:Model"
        ws["AK1"] = "*C:MPN"

        for key, value in result_dic.items():
            i = 2
            for item in value:
                ws[key + str(i)] = item
                i += 1

        wb.save(eel.file_name()())

        eel.view_log_js('ファイルの作成が完了しました')
